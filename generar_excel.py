"""
Genera un archivo Excel (.xlsm) con macros VBA para rastrear
entrenamientos en casa: rutinas por día, series, repeticiones y peso.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import FormulaRule
from copy import copy

# ── Colores y estilos ─────────────────────────────────────────────
AZUL_OSCURO  = "1B2A4A"
AZUL_MEDIO   = "2E5090"
AZUL_CLARO   = "D6E4F0"
VERDE        = "27AE60"
VERDE_CLARO  = "D5F5E3"
NARANJA      = "E67E22"
NARANJA_CL   = "FDEBD0"
GRIS_CLARO   = "F2F2F2"
BLANCO       = "FFFFFF"
ROJO         = "E74C3C"

font_titulo   = Font(name="Calibri", size=16, bold=True, color=BLANCO)
font_subtit   = Font(name="Calibri", size=12, bold=True, color=BLANCO)
font_header   = Font(name="Calibri", size=11, bold=True, color=BLANCO)
font_normal   = Font(name="Calibri", size=11, color="333333")
font_bold     = Font(name="Calibri", size=11, bold=True, color="333333")
font_small    = Font(name="Calibri", size=10, color="666666")
font_btn      = Font(name="Calibri", size=11, bold=True, color=BLANCO)

fill_titulo   = PatternFill("solid", fgColor=AZUL_OSCURO)
fill_header   = PatternFill("solid", fgColor=AZUL_MEDIO)
fill_alt1     = PatternFill("solid", fgColor=BLANCO)
fill_alt2     = PatternFill("solid", fgColor=GRIS_CLARO)
fill_input    = PatternFill("solid", fgColor=AZUL_CLARO)
fill_verde    = PatternFill("solid", fgColor=VERDE)
fill_verde_cl = PatternFill("solid", fgColor=VERDE_CLARO)
fill_naranja  = PatternFill("solid", fgColor=NARANJA)
fill_naranja_cl = PatternFill("solid", fgColor=NARANJA_CL)
fill_rojo     = PatternFill("solid", fgColor=ROJO)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left   = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def set_cell(ws, row, col, value, font=font_normal, fill=None, alignment=align_center, border=thin_border):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    return cell


def merge_and_set(ws, start_row, start_col, end_row, end_col, value, font, fill, alignment=align_center):
    ws.merge_cells(
        start_row=start_row, start_column=start_col,
        end_row=end_row, end_column=end_col,
    )
    cell = ws.cell(row=start_row, column=start_col, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = alignment
    cell.border = thin_border
    # Apply border/fill to all merged cells
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = thin_border
            ws.cell(row=r, column=c).fill = fill


# ── Rutinas predefinidas para casa ────────────────────────────────
RUTINAS = {
    "Lunes - Pecho y Tríceps": [
        "Flexiones clásicas",
        "Flexiones diamante",
        "Flexiones declinadas",
        "Fondos en silla",
        "Flexiones abiertas",
        "Extensión tríceps con mancuerna",
    ],
    "Martes - Espalda y Bíceps": [
        "Remo con mancuernas",
        "Dominadas (o banda elástica)",
        "Remo invertido (mesa)",
        "Curl bíceps con mancuerna",
        "Curl martillo",
        "Superman",
    ],
    "Miércoles - Piernas": [
        "Sentadillas",
        "Sentadilla búlgara",
        "Zancadas",
        "Peso muerto rumano",
        "Elevación de talones",
        "Puente de glúteos",
    ],
    "Jueves - Hombros y Core": [
        "Press militar con mancuernas",
        "Elevaciones laterales",
        "Elevaciones frontales",
        "Plancha frontal (seg)",
        "Plancha lateral (seg)",
        "Crunch abdominal",
    ],
    "Viernes - Full Body": [
        "Burpees",
        "Sentadilla con salto",
        "Flexiones",
        "Remo con mancuernas",
        "Zancadas con salto",
        "Mountain climbers",
    ],
}

DIAS_SEMANA = list(RUTINAS.keys())

# Claves cortas para rangos con nombre (sin espacios ni acentos)
CLAVES_DIA = {
    "Lunes - Pecho y Tríceps":   "Dia_Lunes",
    "Martes - Espalda y Bíceps": "Dia_Martes",
    "Miércoles - Piernas":       "Dia_Miercoles",
    "Jueves - Hombros y Core":   "Dia_Jueves",
    "Viernes - Full Body":       "Dia_Viernes",
}

# Rangos óptimos de repeticiones por ejercicio (min, max)
REPS_RANGES = {
    "Flexiones clásicas": (10, 20),
    "Flexiones diamante": (8, 15),
    "Flexiones declinadas": (8, 15),
    "Fondos en silla": (8, 15),
    "Flexiones abiertas": (10, 20),
    "Extensión tríceps con mancuerna": (10, 15),
    "Remo con mancuernas": (10, 15),
    "Dominadas (o banda elástica)": (5, 12),
    "Remo invertido (mesa)": (8, 15),
    "Curl bíceps con mancuerna": (10, 15),
    "Curl martillo": (10, 15),
    "Superman": (12, 20),
    "Sentadillas": (12, 20),
    "Sentadilla búlgara": (8, 12),
    "Zancadas": (10, 16),
    "Peso muerto rumano": (10, 15),
    "Elevación de talones": (15, 25),
    "Puente de glúteos": (12, 20),
    "Press militar con mancuernas": (8, 12),
    "Elevaciones laterales": (12, 20),
    "Elevaciones frontales": (12, 20),
    "Plancha frontal (seg)": (30, 60),
    "Plancha lateral (seg)": (20, 45),
    "Crunch abdominal": (15, 25),
    "Burpees": (8, 15),
    "Sentadilla con salto": (10, 20),
    "Flexiones": (10, 20),
    "Zancadas con salto": (10, 16),
    "Mountain climbers": (15, 30),
}


def crear_hoja_datos(wb):
    """Hoja oculta con ejercicios en columnas para rangos con nombre."""
    ws = wb.create_sheet("Datos")
    ws.sheet_state = "hidden"

    # Fila 1: claves de día (para MATCH/INDEX desde Registro)
    # Fila 2+: ejercicios de cada día en su columna
    for col_idx, (dia, ejercicios) in enumerate(RUTINAS.items(), start=1):
        clave = CLAVES_DIA[dia]
        # Fila 1: nombre completo del día
        ws.cell(row=1, column=col_idx, value=dia)
        # Fila 2 en adelante: ejercicios
        for row_idx, ej in enumerate(ejercicios, start=2):
            ws.cell(row=row_idx, column=col_idx, value=ej)

        # Crear rango con nombre para esta columna de ejercicios
        col_letter = get_column_letter(col_idx)
        last_row = 1 + len(ejercicios)
        defn = DefinedName(clave, attr_text=f"Datos!${col_letter}$2:${col_letter}${last_row}")
        wb.defined_names.add(defn)

    # Fila 1 también se usa para el lookup: rango "Dias_Lookup"
    col_end = get_column_letter(len(RUTINAS))
    defn_lookup = DefinedName("Dias_Lookup", attr_text=f"Datos!$A$1:${col_end}$1")
    wb.defined_names.add(defn_lookup)

    # ── Tabla de ejercicios con rangos de repeticiones (cols H-J) ──
    # H: Nombre ejercicio, I: Reps Min, J: Reps Max
    ws.cell(row=1, column=8, value="Ejercicio")
    ws.cell(row=1, column=9, value="Reps Min")
    ws.cell(row=1, column=10, value="Reps Max")

    # Deduplicar ejercicios manteniendo orden
    seen = set()
    unique_exercises = []
    for ejercicios in RUTINAS.values():
        for ej in ejercicios:
            if ej not in seen:
                seen.add(ej)
                unique_exercises.append(ej)

    for i, ej in enumerate(unique_exercises, start=2):
        rmin, rmax = REPS_RANGES.get(ej, (8, 15))
        ws.cell(row=i, column=8, value=ej)
        ws.cell(row=i, column=9, value=rmin)
        ws.cell(row=i, column=10, value=rmax)

    last_ex_row = 1 + len(unique_exercises)
    defn_tabla = DefinedName(
        "TablaEjercicios",
        attr_text=f"Datos!$H$1:$J${last_ex_row}")
    wb.defined_names.add(defn_tabla)

    return ws


def crear_hoja_registro(wb):
    """Hoja principal donde se registra cada sesión de entrenamiento."""
    ws = wb.create_sheet("Registro")
    # Move Registro to be the first visible sheet (index 0)
    wb.move_sheet(ws, offset=-wb.sheetnames.index("Registro"))
    ws.sheet_properties.tabColor = AZUL_MEDIO

    # Column widths
    col_widths = {1: 14, 2: 30, 3: 30, 4: 10, 5: 10, 6: 12, 7: 18, 8: 30}
    for c, w in col_widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Título ────────────────────────────────────────────────────
    merge_and_set(ws, 1, 1, 1, 8,
                  "REGISTRO DE ENTRENAMIENTO EN CASA",
                  font_titulo, fill_titulo)

    merge_and_set(ws, 2, 1, 2, 8,
                  "Ingresa tus datos de cada sesión. Usa los botones o escribe directamente.",
                  font_small, PatternFill("solid", fgColor=AZUL_CLARO))

    # ── Zona de entrada rápida ────────────────────────────────────
    merge_and_set(ws, 4, 1, 4, 8,
                  "ENTRADA RÁPIDA",
                  font_subtit, fill_header)

    labels = ["Fecha:", "Día/Rutina:", "Ejercicio:", "Serie #:", "Reps:", "Peso (kg):", "Descanso (seg):", "Notas:"]
    for i, label in enumerate(labels, start=1):
        set_cell(ws, 5, i, label, font_bold, fill_naranja_cl)

    # Input row (row 6)
    for i in range(1, 9):
        set_cell(ws, 6, i, "", font_normal, fill_input)

    # Default date formula
    ws.cell(row=6, column=1).value = '=TODAY()'
    ws.cell(row=6, column=1).number_format = "DD/MM/YYYY"

    # ── Validación desplegable para Día/Rutina ────────────────────
    dv_dia = DataValidation(
        type="list",
        formula1='"' + ",".join(DIAS_SEMANA) + '"',
        allow_blank=True,
    )
    dv_dia.error = "Selecciona un día válido"
    dv_dia.errorTitle = "Día inválido"
    dv_dia.prompt = "Selecciona el día de rutina"
    dv_dia.promptTitle = "Día"
    ws.add_data_validation(dv_dia)
    dv_dia.add(ws["B6"])

    # ── Celda auxiliar oculta (I6) para mapear día → clave de rango ──
    # Fórmula: convierte "Lunes - Pecho y Tríceps" → "Dia_Lunes"
    ws.column_dimensions["I"].width = 0.5  # casi invisible
    ws.column_dimensions["I"].hidden = True
    lookup_formula = (
        '=IF(B6="","",IF(LEFT(B6,5)="Lunes","Dia_Lunes",'
        'IF(LEFT(B6,6)="Martes","Dia_Martes",'
        'IF(LEFT(B6,4)="Mi' + chr(233) + 'r","Dia_Miercoles",'
        'IF(LEFT(B6,6)="Jueves","Dia_Jueves",'
        'IF(LEFT(B6,7)="Viernes","Dia_Viernes",""))))))'
    )
    ws.cell(row=6, column=9, value=lookup_formula)
    ws.cell(row=6, column=9).font = Font(color=BLANCO, size=1)

    # ── Validación desplegable dinámica para Ejercicio (C6) ──────
    dv_ejercicio = DataValidation(
        type="list",
        formula1='=INDIRECT(I6)',
        allow_blank=True,
    )
    dv_ejercicio.error = "Selecciona un ejercicio de la rutina"
    dv_ejercicio.errorTitle = "Ejercicio inválido"
    dv_ejercicio.prompt = "Selecciona un ejercicio de la rutina del día"
    dv_ejercicio.promptTitle = "Ejercicio"
    dv_ejercicio.showErrorMessage = False  # permitir escribir ejercicios custom
    ws.add_data_validation(dv_ejercicio)
    dv_ejercicio.add(ws["C6"])

    # ── Validación numérica para Serie # (D6) ─────────────────────
    dv_serie = DataValidation(
        type="list",
        formula1='"1,2,3,4,5,6,7,8,9,10"',
        allow_blank=True,
    )
    dv_serie.prompt = "N\u00famero de serie (1, 2, 3...)"
    dv_serie.promptTitle = "Serie #"
    ws.add_data_validation(dv_serie)
    dv_serie.add(ws["D6"])

    # ── Indicador de rango de reps en fila 7 ──────────────────────
    status_formula = (
        '=IF(OR(C6="",E6=""),"",IF(ISERROR(VLOOKUP(C6,TablaEjercicios,2,FALSE)),'
        '"",IF(E6>VLOOKUP(C6,TablaEjercicios,3,FALSE),'
        '"\u2B06 SUPERA RANGO - SUBE PESO",'
        'IF(E6<VLOOKUP(C6,TablaEjercicios,2,FALSE),'
        '"\u2B07 BAJO RANGO - BAJA PESO",'
        '"\u2714 EN RANGO \u00d3PTIMO"))))'
    )
    merge_and_set(ws, 7, 4, 7, 7, "", font_small,
                  PatternFill("solid", fgColor=AZUL_CLARO))
    cell_status = ws.cell(row=7, column=4)
    cell_status.value = status_formula
    cell_status.font = Font(name="Calibri", size=10, bold=True, color="333333")
    cell_status.alignment = align_center

    # Mostrar rango objetivo del ejercicio seleccionado
    range_formula = (
        '=IF(C6="","",IF(ISERROR(VLOOKUP(C6,TablaEjercicios,2,FALSE)),'
        '"","Rango: "&VLOOKUP(C6,TablaEjercicios,2,FALSE)'
        '&" - "&VLOOKUP(C6,TablaEjercicios,3,FALSE)&" reps"))'
    )
    merge_and_set(ws, 7, 1, 7, 3, "", font_small,
                  PatternFill("solid", fgColor=AZUL_CLARO))
    cell_range = ws.cell(row=7, column=1)
    cell_range.value = range_formula
    cell_range.font = Font(name="Calibri", size=10, italic=True, color="555555")
    cell_range.alignment = align_center

    # ── Botones de acción (celdas con texto + macro asignada via VBA) ──
    merge_and_set(ws, 8, 2, 8, 3,
                  "▶ REGISTRAR ENTRADA",
                  font_btn, fill_verde)
    merge_and_set(ws, 8, 5, 8, 6,
                  "✕ LIMPIAR CAMPOS",
                  font_btn, fill_naranja)
    merge_and_set(ws, 8, 7, 8, 8,
                  "⟳ DESHACER ÚLTIMO",
                  font_btn, fill_rojo)

    # ── Encabezados del historial ─────────────────────────────────
    merge_and_set(ws, 10, 1, 10, 8,
                  "HISTORIAL DE ENTRENAMIENTOS",
                  font_subtit, fill_titulo)

    headers = ["Fecha", "D\u00eda / Rutina", "Ejercicio", "Serie #",
               "Reps", "Peso (kg)", "Descanso (s)", "Notas"]
    for i, h in enumerate(headers, start=1):
        set_cell(ws, 11, i, h, font_header, fill_header)

    # Rows 12-200 pre-formatted for data
    for r in range(12, 201):
        fill = fill_alt1 if r % 2 == 0 else fill_alt2
        for c in range(1, 9):
            cell = set_cell(ws, r, c, "", font_normal, fill)
            if c == 1:
                cell.number_format = "DD/MM/YYYY"

    # ── Formato condicional: alertas de rango de reps ─────────────
    # Rojo: reps superan el máximo → hay que subir peso
    fill_alert_high = PatternFill(
        start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    font_alert_high = Font(name="Calibri", size=11, bold=True, color="CC0000")
    rule_high = FormulaRule(
        formula=['AND(E12<>"",NOT(ISERROR(VLOOKUP($C12,TablaEjercicios,3,FALSE))),'
                 'E12>VLOOKUP($C12,TablaEjercicios,3,FALSE))'],
        fill=fill_alert_high,
        font=font_alert_high,
    )
    ws.conditional_formatting.add("E12:E200", rule_high)

    # Amarillo: reps por debajo del mínimo → considerar bajar peso
    fill_alert_low = PatternFill(
        start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    font_alert_low = Font(name="Calibri", size=11, bold=True, color="856404")
    rule_low = FormulaRule(
        formula=['AND(E12<>"",NOT(ISERROR(VLOOKUP($C12,TablaEjercicios,2,FALSE))),'
                 'E12<VLOOKUP($C12,TablaEjercicios,2,FALSE))'],
        fill=fill_alert_low,
        font=font_alert_low,
    )
    ws.conditional_formatting.add("E12:E200", rule_low)

    # Verde: reps dentro del rango óptimo
    fill_ok = PatternFill(
        start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    font_ok = Font(name="Calibri", size=11, color="155724")
    rule_ok = FormulaRule(
        formula=['AND(E12<>"",NOT(ISERROR(VLOOKUP($C12,TablaEjercicios,2,FALSE))),'
                 'E12>=VLOOKUP($C12,TablaEjercicios,2,FALSE),'
                 'E12<=VLOOKUP($C12,TablaEjercicios,3,FALSE))'],
        fill=fill_ok,
        font=font_ok,
    )
    ws.conditional_formatting.add("E12:E200", rule_ok)

    # ── Freeze panes ──────────────────────────────────────────────
    ws.freeze_panes = "A12"

    # ── Auto-filter ───────────────────────────────────────────────
    ws.auto_filter.ref = "A11:H200"

    return ws


def crear_hoja_rutinas(wb):
    """Hoja con las rutinas predefinidas por día."""
    ws = wb.create_sheet("Rutinas")
    ws.sheet_properties.tabColor = VERDE

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14

    merge_and_set(ws, 1, 1, 1, 7,
                  "RUTINAS SEMANALES - ENTRENAMIENTO EN CASA",
                  font_titulo, fill_titulo)

    merge_and_set(ws, 2, 1, 2, 7,
                  "Personaliza tus rutinas aqu\u00ed. Ajusta los rangos de reps "
                  "para recibir alertas autom\u00e1ticas.",
                  font_small, PatternFill("solid", fgColor=VERDE_CLARO))

    row = 4
    for dia, ejercicios in RUTINAS.items():
        merge_and_set(ws, row, 1, row, 7, dia, font_subtit, fill_header)
        row += 1

        sub_headers = ["#", "Ejercicio", "Grupo Muscular",
                       "Series Obj.", "Reps Min", "Reps Max", "Peso Obj. (kg)"]
        for i, h in enumerate(sub_headers, start=1):
            set_cell(ws, row, i, h, font_header, fill_verde)
        row += 1

        for idx, ej in enumerate(ejercicios, start=1):
            fill = fill_alt1 if idx % 2 == 0 else fill_alt2
            rmin, rmax = REPS_RANGES.get(ej, (8, 15))
            set_cell(ws, row, 1, idx, font_normal, fill)
            set_cell(ws, row, 2, ej, font_normal, fill, align_left)
            set_cell(ws, row, 3, "", font_normal, fill)       # Grupo muscular
            set_cell(ws, row, 4, "", font_normal, fill)       # Series obj
            set_cell(ws, row, 5, rmin, font_normal, fill)     # Reps Min
            set_cell(ws, row, 6, rmax, font_normal, fill)     # Reps Max
            set_cell(ws, row, 7, "", font_normal, fill)       # Peso obj
            row += 1

        row += 1  # espacio entre días

    return ws


def crear_hoja_dashboard(wb):
    """Dashboard visual con KPIs, gráficos e insights de progreso."""
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = NARANJA

    # ── Column widths ──────────────────────────────────────────────
    ws.column_dimensions["A"].width = 2       # spacer
    ws.column_dimensions["B"].width = 26      # labels
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 2       # spacer
    for letter in ["H", "I", "J", "K", "L"]:
        ws.column_dimensions[letter].width = 12

    # Fuentes especiales para KPIs
    font_kpi_label = Font(name="Calibri", size=9, bold=True, color=BLANCO)
    font_kpi_value = Font(name="Calibri", size=22, bold=True, color=BLANCO)
    font_section   = Font(name="Calibri", size=12, bold=True, color=BLANCO)

    # ══════════════════════════════════════════════════════════════
    # TÍTULO
    # ══════════════════════════════════════════════════════════════
    merge_and_set(ws, 1, 1, 1, 12,
                  "DASHBOARD DE ENTRENAMIENTO",
                  font_titulo, fill_titulo)
    merge_and_set(ws, 2, 1, 2, 12,
                  "Resumen visual de tu progreso. Los datos se actualizan "
                  "autom\u00e1ticamente desde la hoja Registro.",
                  font_small, PatternFill("solid", fgColor=NARANJA_CL))

    # ══════════════════════════════════════════════════════════════
    # KPI CARDS  (row 4-7)
    # ══════════════════════════════════════════════════════════════
    merge_and_set(ws, 4, 1, 4, 12,
                  "M\u00c9TRICAS CLAVE", font_section, fill_header)

    kpis = [
        ("Total Sesiones",
         '=COUNTA(Registro!A12:A200)',
         None, VERDE),
        ("\u00daltimo Entreno",
         '=IF(COUNTA(Registro!A12:A200)>0,MAX(Registro!A12:A200),"-")',
         "DD/MM/YYYY", AZUL_MEDIO),
        ("Peso M\u00e1x (kg)",
         '=IF(MAX(Registro!F12:F200)>0,MAX(Registro!F12:F200),"-")',
         "0.0", NARANJA),
        ("Prom. Reps",
         '=IFERROR(ROUND(AVERAGE(Registro!E12:E200),1),"-")',
         "0.0", AZUL_OSCURO),
        ("Total Sets",
         '=IFERROR(COUNTA(Registro!D12:D200),"-")',
         "#,##0", ROJO),
    ]

    col_starts = [2, 4, 6, 8, 10]
    for idx, (label, formula, num_fmt, color) in enumerate(kpis):
        c = col_starts[idx]
        fill_kpi = PatternFill("solid", fgColor=color)
        # Label row
        merge_and_set(ws, 5, c, 5, c + 1, label, font_kpi_label, fill_kpi)
        # Value row (2 rows tall)
        merge_and_set(ws, 6, c, 7, c + 1, "", font_kpi_value, fill_kpi)
        cell = ws.cell(row=6, column=c)
        cell.value = formula
        cell.font = font_kpi_value
        cell.alignment = align_center
        if num_fmt:
            cell.number_format = num_fmt

    # ══════════════════════════════════════════════════════════════
    # VOLUMEN POR DÍA  (rows 9-15)  — left side
    # ══════════════════════════════════════════════════════════════
    merge_and_set(ws, 9, 2, 9, 6,
                  "VOLUMEN POR D\u00cdA DE RUTINA", font_section, fill_header)

    vol_headers = ["D\u00eda", "Sets", "Total Reps", "Reps Prom.", "Peso Prom."]
    for i, h in enumerate(vol_headers):
        set_cell(ws, 10, 2 + i, h, font_header, fill_verde)

    for i, dia in enumerate(DIAS_SEMANA):
        r = 11 + i
        fill = fill_alt1 if i % 2 == 0 else fill_alt2
        short = dia.split(" - ")[0]
        set_cell(ws, r, 2, short, font_bold, fill, align_left)
        # Sets = count of rows for this day
        set_cell(ws, r, 3,
                 f'=COUNTIF(Registro!B$12:B$200,"{dia}")',
                 font_normal, fill)
        # Total reps
        set_cell(ws, r, 4,
                 f'=SUMIF(Registro!B$12:B$200,"{dia}",Registro!E$12:E$200)',
                 font_normal, fill)
        # Avg reps per set
        set_cell(ws, r, 5,
                 f'=IFERROR(AVERAGEIF(Registro!B$12:B$200,"{dia}",'
                 f'Registro!E$12:E$200),0)',
                 font_normal, fill)
        set_cell(ws, r, 6,
                 f'=IFERROR(AVERAGEIF(Registro!B$12:B$200,"{dia}",'
                 f'Registro!F$12:F$200),0)',
                 font_normal, fill)

    # ── Gráfico de barras: Sesiones por Día ────────────────────────
    chart_bar = BarChart()
    chart_bar.type = "col"
    chart_bar.style = 10
    chart_bar.title = "Sesiones por D\u00eda"
    chart_bar.y_axis.title = "Sesiones"
    chart_bar.x_axis.delete = False
    chart_bar.legend = None

    data_bar = Reference(ws, min_col=3, min_row=10, max_row=15)
    cats_bar = Reference(ws, min_col=2, min_row=11, max_row=15)
    chart_bar.add_data(data_bar, titles_from_data=True)
    chart_bar.set_categories(cats_bar)
    chart_bar.width = 20
    chart_bar.height = 12

    series_bar = chart_bar.series[0]
    series_bar.graphicalProperties.solidFill = AZUL_MEDIO

    ws.add_chart(chart_bar, "H9")

    # ══════════════════════════════════════════════════════════════
    # INSIGHTS  (rows 17-24)  — left side
    # ══════════════════════════════════════════════════════════════
    merge_and_set(ws, 17, 2, 17, 6,
                  "INSIGHTS", font_section, fill_titulo)

    insights = [
        ("D\u00edas sin entrenar",
         '=IF(COUNTA(Registro!A12:A200)>0,'
         'TODAY()-MAX(Registro!A12:A200),"-")'),
        ("Sesiones esta semana",
         '=COUNTIFS(Registro!A12:A200,">="&(TODAY()-WEEKDAY(TODAY(),2)+1),'
         'Registro!A12:A200,"<="&TODAY())'),
        ("Sesiones este mes",
         '=COUNTIFS(Registro!A12:A200,">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'
         'Registro!A12:A200,"<="&TODAY())'),
        ("Ejercicio m\u00e1s frecuente",
         '=IFERROR(INDEX(Registro!C12:C200,MATCH(MAX(COUNTIF(Registro!C12:C200,'
         'Registro!C12:C200)),COUNTIF(Registro!C12:C200,Registro!C12:C200),0)),"-")'),
        ("Total reps acumuladas",
         '=IFERROR(SUM(Registro!E12:E200),"-")'),
        ("Reps max en 1 set",
         '=IFERROR(MAX(Registro!E12:E200),"-")'),
    ]

    for i, (label, formula) in enumerate(insights):
        r = 18 + i
        fill = fill_alt1 if i % 2 == 0 else fill_alt2
        set_cell(ws, r, 2, label, font_bold, fill, align_left)
        merge_and_set(ws, r, 3, r, 6, "", font_normal, fill)
        cell = ws.cell(row=r, column=3)
        cell.value = formula
        cell.font = font_bold
        cell.alignment = align_center

    # ── Gráfico circular: Distribución ─────────────────────────────
    chart_pie = PieChart()
    chart_pie.title = "Distribuci\u00f3n por Rutina"
    chart_pie.style = 10

    data_pie = Reference(ws, min_col=3, min_row=10, max_row=15)
    cats_pie = Reference(ws, min_col=2, min_row=11, max_row=15)
    chart_pie.add_data(data_pie, titles_from_data=True)
    chart_pie.set_categories(cats_pie)
    chart_pie.width = 20
    chart_pie.height = 12

    chart_pie.dataLabels = DataLabelList()
    chart_pie.dataLabels.showPercent = True
    chart_pie.dataLabels.showCatName = True
    chart_pie.dataLabels.showVal = False

    colors_pie = [AZUL_MEDIO, VERDE, NARANJA, ROJO, AZUL_OSCURO]
    for i, color in enumerate(colors_pie):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        chart_pie.series[0].data_points.append(pt)

    ws.add_chart(chart_pie, "H22")

    # ══════════════════════════════════════════════════════════════
    # ALERTAS DE PROGRESO  (rows 26-30)  — left side
    # ══════════════════════════════════════════════════════════════
    fill_alert_section = PatternFill("solid", fgColor=ROJO)
    merge_and_set(ws, 26, 2, 26, 6,
                  "\u26a0 ALERTAS DE PROGRESO", font_section, fill_alert_section)

    alertas = [
        ("Registros SOBRE rango (subir peso)",
         '=IFERROR(SUMPRODUCT((Registro!E12:E200<>"")'
         '*NOT(ISERROR(VLOOKUP(Registro!C12:C200,TablaEjercicios,3,FALSE)))'
         '*(Registro!E12:E200>VLOOKUP(Registro!C12:C200,TablaEjercicios,3,FALSE))),0)'),
        ("Registros BAJO rango (bajar peso)",
         '=IFERROR(SUMPRODUCT((Registro!E12:E200<>"")'
         '*NOT(ISERROR(VLOOKUP(Registro!C12:C200,TablaEjercicios,2,FALSE)))'
         '*(Registro!E12:E200<VLOOKUP(Registro!C12:C200,TablaEjercicios,2,FALSE))),0)'),
        ("% en rango \u00f3ptimo",
         '=IFERROR(TEXT(1-((SUMPRODUCT((Registro!E12:E200<>"")'
         '*NOT(ISERROR(VLOOKUP(Registro!C12:C200,TablaEjercicios,3,FALSE)))'
         '*(Registro!E12:E200>VLOOKUP(Registro!C12:C200,TablaEjercicios,3,FALSE)))'
         '+SUMPRODUCT((Registro!E12:E200<>"")'
         '*NOT(ISERROR(VLOOKUP(Registro!C12:C200,TablaEjercicios,2,FALSE)))'
         '*(Registro!E12:E200<VLOOKUP(Registro!C12:C200,TablaEjercicios,2,FALSE))))'
         '/MAX(COUNTA(Registro!E12:E200),1)),"0%"),"-")'),
    ]

    for i, (label, formula) in enumerate(alertas):
        r = 27 + i
        fill = fill_alt1 if i % 2 == 0 else fill_alt2
        set_cell(ws, r, 2, label, font_bold, fill, align_left)
        merge_and_set(ws, r, 3, r, 6, "", font_normal, fill)
        cell = ws.cell(row=r, column=3)
        cell.value = formula
        cell.font = font_bold
        cell.alignment = align_center

    # ══════════════════════════════════════════════════════════════
    # ÚLTIMOS 10 REGISTROS  (rows 31-42)
    # ══════════════════════════════════════════════════════════════
    merge_and_set(ws, 31, 2, 31, 6,
                  "\u00daLTIMOS 10 REGISTROS", font_section, fill_titulo)

    last_headers = ["Fecha", "Rutina", "Ejercicio", "Reps", "Peso (kg)"]
    for i, h in enumerate(last_headers):
        set_cell(ws, 32, 2 + i, h, font_header, fill_verde)

    for i in range(10):
        r = 33 + i
        fill = fill_alt1 if i % 2 == 0 else fill_alt2
        for c in range(2, 7):
            set_cell(ws, r, c, "", font_normal, fill)
        cols_map = {2: "A", 3: "B", 4: "C", 5: "E", 6: "F"}
        for c, col_letter in cols_map.items():
            formula = (f'=IFERROR(INDEX(Registro!{col_letter}$12:'
                       f'{col_letter}$200,COUNTA(Registro!A$12:A$200)-{i}),"")')
            ws.cell(row=r, column=c).value = formula
            if c == 2:
                ws.cell(row=r, column=c).number_format = "DD/MM/YYYY"

    # ══════════════════════════════════════════════════════════════
    # GRÁFICO DE LÍNEA: Peso por sesión  (from Registro)
    # ══════════════════════════════════════════════════════════════
    ws_reg = wb["Registro"]
    chart_line = LineChart()
    chart_line.title = "Peso Levantado por Sesi\u00f3n"
    chart_line.y_axis.title = "Peso (kg)"
    chart_line.x_axis.title = "Entrada"
    chart_line.style = 10
    chart_line.legend = None
    chart_line.width = 20
    chart_line.height = 12

    data_line = Reference(ws_reg, min_col=6, min_row=11, max_row=200)
    chart_line.add_data(data_line, titles_from_data=True)

    series_line = chart_line.series[0]
    series_line.graphicalProperties.line.solidFill = VERDE
    series_line.graphicalProperties.line.width = 22000
    series_line.smooth = True

    ws.add_chart(chart_line, "H41")

    # ── Freeze panes ───────────────────────────────────────────────
    ws.freeze_panes = "A4"

    return ws


def crear_hoja_instrucciones(wb):
    """Hoja con instrucciones de uso."""
    ws = wb.create_sheet("Instrucciones")
    ws.sheet_properties.tabColor = "95A5A6"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 80

    merge_and_set(ws, 1, 1, 1, 2,
                  "INSTRUCCIONES DE USO",
                  font_titulo, fill_titulo)

    instrucciones = [
        ("HOJA 'REGISTRO'", [
            "Cada fila = UN SET de un ejercicio (Serie 1: 15 reps, Serie 2: 12 reps, etc.).",
            "1. La FECHA se llena autom\u00e1ticamente con el d\u00eda de hoy.",
            "2. Selecciona el D\u00cdA/RUTINA y luego el EJERCICIO del desplegable din\u00e1mico.",
            "3. Ingresa el SERIE # (1, 2, 3...), REPS y PESO de ese set.",
            "4. Al registrar, el Serie # sube autom\u00e1ticamente para el siguiente set.",
            "5. La fila 7 muestra si tus reps est\u00e1n EN RANGO, SOBRE o BAJO el \u00f3ptimo.",
            "6. En el historial: VERDE = en rango, ROJO = sube peso, AMARILLO = baja peso.",
            "7. Haz clic en '\u25b6 REGISTRAR ENTRADA' para guardar el set.",
            "8. Usa '\u2715 LIMPIAR CAMPOS' o '\u27f3 DESHACER \u00daLTIMO' seg\u00fan necesites.",
        ]),
        ("HOJA 'RUTINAS'", [
            "Rutinas predefinidas para 5 d\u00edas con rangos de repeticiones \u00f3ptimas.",
            "Columnas Reps Min y Reps Max definen el rango ideal por ejercicio.",
            "Si superas Reps Max consistentemente, es hora de SUBIR PESO.",
            "Puedes editar los rangos para personalizar tus objetivos.",
        ]),
        ("HOJA 'DASHBOARD'", [
            "Dashboard visual con KPIs, gr\u00e1ficos e insights de tu progreso.",
            "5 tarjetas de m\u00e9tricas clave: sesiones, \u00faltimo entreno, peso m\u00e1x, etc.",
            "Gr\u00e1fico de barras: sesiones por d\u00eda de rutina.",
            "Gr\u00e1fico circular: distribuci\u00f3n porcentual de entrenamientos.",
            "Gr\u00e1fico de l\u00ednea: progresi\u00f3n del peso levantado por sesi\u00f3n.",
            "Secci\u00f3n de insights: d\u00edas sin entrenar, sesiones semanales/mensuales.",
            "Los \u00faltimos 10 registros se muestran al final.",
        ]),
        ("SISTEMA DE ALERTAS", [
            "Cada ejercicio tiene un rango \u00f3ptimo de repeticiones (Reps Min - Reps Max).",
            "VERDE en la columna Reps = est\u00e1s dentro del rango \u00f3ptimo.",
            "ROJO = superaste el m\u00e1ximo \u2192 SUBE EL PESO en tu pr\u00f3ximo entrenamiento.",
            "AMARILLO = est\u00e1s por debajo del m\u00ednimo \u2192 considera BAJAR PESO.",
            "El Dashboard muestra cu\u00e1ntos registros est\u00e1n fuera de rango.",
        ]),
        ("TIPS", [
            "Registra CADA SET por separado: Serie 1 (15 reps), Serie 2 (12 reps), etc.",
            "El Serie # sube solo al registrar \u2014 perfecto para m\u00faltiples sets seguidos.",
            "Revisa el Dashboard regularmente para ver alertas y tu progreso.",
            "\u00a1La constancia es la clave! Intenta entrenar al menos 3-4 d\u00edas por semana.",
        ]),
    ]

    row = 3
    for titulo, lineas in instrucciones:
        merge_and_set(ws, row, 1, row, 2, titulo, font_subtit, fill_header)
        row += 1
        for linea in lineas:
            set_cell(ws, row, 1, "•", font_normal, fill_alt1)
            set_cell(ws, row, 2, linea, font_normal, fill_alt1, align_left)
            row += 1
        row += 1

    return ws


# ── Código VBA para los botones ───────────────────────────────────
VBA_CODE = '''
Attribute VB_Name = "ModEntrenamiento"

' ============================================================
' MACRO: Registrar una entrada desde la zona de entrada rápida
' ============================================================
Sub RegistrarEntrada()
    Dim wsReg As Worksheet
    Set wsReg = ThisWorkbook.Sheets("Registro")

    ' Buscar la siguiente fila vacía en el historial (desde fila 12)
    Dim nextRow As Long
    nextRow = 12
    Do While wsReg.Cells(nextRow, 1).Value <> ""
        nextRow = nextRow + 1
        If nextRow > 200 Then
            MsgBox "El historial esta lleno (200 registros). " & _
                   "Copia los datos a otro archivo para continuar.", _
                   vbExclamation, "Historial lleno"
            Exit Sub
        End If
    Loop

    ' Validar que haya al menos un ejercicio
    If Trim(wsReg.Cells(6, 3).Value) = "" Then
        MsgBox "Por favor ingresa al menos el nombre del ejercicio.", _
               vbExclamation, "Falta informacion"
        Exit Sub
    End If

    ' Copiar datos de la fila 6 (entrada) a la fila del historial
    Dim col As Integer
    For col = 1 To 8
        wsReg.Cells(nextRow, col).Value = wsReg.Cells(6, col).Value
    Next col

    ' Formatear la fecha
    wsReg.Cells(nextRow, 1).NumberFormat = "DD/MM/YYYY"

    ' Auto-incrementar Serie # y limpiar reps/peso (mantener ejercicio y dia)
    Dim serieActual As Variant
    serieActual = wsReg.Cells(6, 4).Value
    If IsNumeric(serieActual) Then
        wsReg.Cells(6, 4).Value = CLng(serieActual) + 1
    Else
        wsReg.Cells(6, 4).Value = 2
    End If

    ' Limpiar solo reps, peso, descanso y notas
    wsReg.Cells(6, 5).Value = ""  ' Reps
    wsReg.Cells(6, 6).Value = ""  ' Peso
    wsReg.Cells(6, 7).Value = ""  ' Descanso
    wsReg.Cells(6, 8).Value = ""  ' Notas

    MsgBox "Serie " & wsReg.Cells(nextRow, 4).Value & " de " & _
           wsReg.Cells(nextRow, 3).Value & " registrada." & vbCrLf & _
           "Siguiente: Serie " & wsReg.Cells(6, 4).Value, _
           vbInformation, "Registrado"
End Sub


' ============================================================
' MACRO: Limpiar todos los campos de entrada rápida
' ============================================================
Sub LimpiarCampos()
    Dim wsReg As Worksheet
    Set wsReg = ThisWorkbook.Sheets("Registro")

    wsReg.Cells(6, 1).Value = Date
    wsReg.Cells(6, 1).NumberFormat = "DD/MM/YYYY"
    wsReg.Cells(6, 2).Value = ""
    wsReg.Cells(6, 3).Value = ""
    wsReg.Cells(6, 4).Value = ""
    wsReg.Cells(6, 5).Value = ""
    wsReg.Cells(6, 6).Value = ""
    wsReg.Cells(6, 7).Value = ""
    wsReg.Cells(6, 8).Value = ""

    MsgBox "Campos limpiados.", vbInformation, "Listo"
End Sub


' ============================================================
' MACRO: Deshacer el último registro (eliminar última fila con datos)
' ============================================================
Sub DeshacerUltimo()
    Dim wsReg As Worksheet
    Set wsReg = ThisWorkbook.Sheets("Registro")

    ' Encontrar la última fila con datos
    Dim lastRow As Long
    lastRow = 11
    Do While wsReg.Cells(lastRow + 1, 1).Value <> ""
        lastRow = lastRow + 1
    Loop

    If lastRow < 12 Then
        MsgBox "No hay registros para deshacer.", _
               vbExclamation, "Sin registros"
        Exit Sub
    End If

    Dim resp As VbMsgBoxResult
    resp = MsgBox("Eliminar el ultimo registro?" & vbCrLf & _
                  "Fecha: " & wsReg.Cells(lastRow, 1).Value & vbCrLf & _
                  "Ejercicio: " & wsReg.Cells(lastRow, 3).Value, _
                  vbYesNo + vbQuestion, "Confirmar")

    If resp = vbYes Then
        Dim col As Integer
        For col = 1 To 8
            wsReg.Cells(lastRow, col).Value = ""
        Next col
        MsgBox "Ultimo registro eliminado.", vbInformation, "Deshecho"
    End If
End Sub


' ============================================================
' MACRO: Cargar ejercicios de la rutina seleccionada
' ============================================================
Sub CargarEjercicios()
    Dim wsReg As Worksheet, wsRut As Worksheet
    Set wsReg = ThisWorkbook.Sheets("Registro")
    Set wsRut = ThisWorkbook.Sheets("Rutinas")

    Dim diaSeleccionado As String
    diaSeleccionado = Trim(wsReg.Cells(6, 2).Value)

    If diaSeleccionado = "" Then
        MsgBox "Primero selecciona un Dia/Rutina en la celda B6.", _
               vbExclamation, "Selecciona un dia"
        Exit Sub
    End If

    ' Buscar el día en la hoja Rutinas
    Dim r As Long
    Dim found As Boolean
    found = False

    For r = 1 To 100
        If InStr(1, wsRut.Cells(r, 1).Value, diaSeleccionado, vbTextCompare) > 0 Then
            found = True
            Exit For
        End If
    Next r

    If Not found Then
        MsgBox "No se encontro la rutina: " & diaSeleccionado, _
               vbExclamation, "No encontrado"
        Exit Sub
    End If

    ' Leer ejercicios (están 2 filas abajo del título del día)
    Dim ejercicios As String
    Dim startRow As Long
    startRow = r + 2  ' Saltar fila de sub-headers

    ejercicios = "Ejercicios para " & diaSeleccionado & ":" & vbCrLf & vbCrLf

    Dim ej As Long
    ej = 0
    Do While wsRut.Cells(startRow + ej, 2).Value <> "" And ej < 20
        ejercicios = ejercicios & (ej + 1) & ". " & wsRut.Cells(startRow + ej, 2).Value & vbCrLf
        ej = ej + 1
    Loop

    MsgBox ejercicios, vbInformation, "Rutina del dia"
End Sub
'''


def main():
    wb = openpyxl.Workbook()

    # Eliminar la hoja por defecto
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Crear hojas
    crear_hoja_datos(wb)        # primero: crea rangos con nombre
    crear_hoja_registro(wb)
    crear_hoja_rutinas(wb)
    crear_hoja_dashboard(wb)
    crear_hoja_instrucciones(wb)

    # Mover Instrucciones al principio? No, dejarlo al final.
    # Asegurar que Registro es la hoja activa
    wb.active = wb.sheetnames.index("Registro")

    # ── Guardar como .xlsm con VBA ───────────────────────────────
    # openpyxl no soporta VBA nativamente en archivos nuevos.
    # Guardamos como .xlsx y creamos un archivo .bas con las macros.
    output_xlsx = "/home/user/gym/Entrenamiento_Casa.xlsx"
    output_vba  = "/home/user/gym/macros_entrenamiento.bas"

    wb.save(output_xlsx)

    # Guardar el código VBA como archivo .bas importable
    with open(output_vba, "w", encoding="utf-8") as f:
        f.write(VBA_CODE)

    print(f"Archivo Excel creado: {output_xlsx}")
    print(f"Archivo de macros VBA: {output_vba}")
    print()
    print("INSTRUCCIONES PARA ACTIVAR LAS MACROS:")
    print("=" * 50)
    print("1. Abre el archivo .xlsx en Excel")
    print("2. Presiona Alt+F11 para abrir el Editor de VBA")
    print("3. Ve a Archivo > Importar archivo...")
    print("4. Selecciona 'macros_entrenamiento.bas'")
    print("5. Cierra el editor VBA")
    print("6. Guarda el archivo como .xlsm (Libro habilitado para macros)")
    print()
    print("ASIGNAR MACROS A LOS BOTONES:")
    print("=" * 50)
    print("1. Clic derecho sobre '▶ REGISTRAR ENTRADA' > Asignar macro > RegistrarEntrada")
    print("2. Clic derecho sobre '✕ LIMPIAR CAMPOS' > Asignar macro > LimpiarCampos")
    print("3. Clic derecho sobre '⟳ DESHACER ÚLTIMO' > Asignar macro > DeshacerUltimo")
    print()
    print("ALTERNATIVA RÁPIDA (sin importar .bas):")
    print("=" * 50)
    print("En la hoja de Registro, usa Alt+F8 para ver las macros disponibles")
    print("y ejecutarlas manualmente.")


if __name__ == "__main__":
    main()
